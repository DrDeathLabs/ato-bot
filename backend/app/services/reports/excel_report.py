"""Excel control matrix report generator."""
from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from app.core.config import get_settings
from app.core.database import AsyncSessionLocal
from app.models.orm import Assessment, ControlFinding, Project

settings = get_settings()

STATUS_COLORS = {
    "compliant": "C6EFCE",
    "partially_compliant": "FFEB9C",
    "non_compliant": "FFC7CE",
    "not_applicable": "D9D9D9",
    "not_reviewed": "FFFFFF",
}
STATUS_LABELS = {
    "compliant": "Compliant",
    "partially_compliant": "Partially Compliant",
    "non_compliant": "Non-Compliant",
    "not_applicable": "Not Applicable",
    "not_reviewed": "Not Reviewed",
}

# NIST SP 800-53A Rev 5 formal determination labels for SAR/ATO package output
NIST_DETERMINATION_LABELS = {
    "compliant": "Satisfied",
    "partially_compliant": "Other Than Satisfied",
    "non_compliant": "Other Than Satisfied",
    "not_applicable": "Not Applicable",
    "not_reviewed": "Not Reviewed",
}
NIST_DETERMINATION_COLORS = {
    "Satisfied": "C6EFCE",
    "Other Than Satisfied": "FFC7CE",
    "Not Applicable": "D9D9D9",
    "Not Reviewed": "FFFFFF",
}


async def generate_excel(assessment_id: int) -> str:
    async with AsyncSessionLocal() as db:
        result = await db.execute(select(Assessment).where(Assessment.id == assessment_id))
        assessment: Assessment = result.scalar_one()

        proj_result = await db.execute(select(Project).where(Project.id == assessment.project_id))
        project: Project = proj_result.scalar_one()

        findings_result = await db.execute(
            select(ControlFinding)
            .where(ControlFinding.assessment_id == assessment_id)
            .order_by(ControlFinding.control_family, ControlFinding.control_id)
        )
        findings = findings_result.scalars().all()

    wb = Workbook()

    # ── Summary Sheet ──────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, project, assessment, findings)

    # ── Control Matrix Sheet ───────────────────────────────────────────────────
    ws_matrix = wb.create_sheet("Control Matrix")
    _write_control_matrix(ws_matrix, findings)

    # ── Findings by Family ─────────────────────────────────────────────────────
    families = sorted({f.control_family for f in findings})
    for family in families:
        ws_fam = wb.create_sheet(family)
        fam_findings = [f for f in findings if f.control_family == family]
        _write_family_sheet(ws_fam, family, fam_findings)

    # Save
    output_path = Path(settings.output_dir) / f"assessment_{assessment_id}_controls.xlsx"
    os.makedirs(settings.output_dir, exist_ok=True)
    wb.save(str(output_path))
    return str(output_path)


def _header_style(ws, row, values: list, bold: bool = True, bg_color: str = "1F4E79"):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=bold, color="FFFFFF" if bg_color != "FFFFFF" else "000000")
        cell.fill = PatternFill("solid", fgColor=bg_color)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(
            bottom=Side(style="thin"), right=Side(style="thin")
        )


def _write_summary(ws, project, assessment, findings):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    ws["A1"] = "ATO Assessment Summary"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:B1")

    # NIST 800-53A formal determination counts
    satisfied = sum(1 for f in findings if f.status == "compliant")
    ots = sum(1 for f in findings if f.status in ("partially_compliant", "non_compliant"))
    synthesized_count = sum(1 for f in findings if getattr(f, "synthesized_narrative", False))
    challenged_count = sum(1 for f in findings if getattr(f, "llm_challenge_note", None))

    rows = [
        ("System Name", project.name),
        ("Impact Baseline", project.impact_baseline.upper()),
        ("Assessment Date", assessment.completed_at.strftime("%Y-%m-%d") if assessment.completed_at else ""),
        ("LLM Provider", f"{assessment.llm_provider} / {assessment.llm_model}"),
        ("", ""),
        ("── Working View (Internal) ──", ""),
        ("Total Controls", len(findings)),
        ("Compliant", sum(1 for f in findings if f.status == "compliant")),
        ("Partially Compliant", sum(1 for f in findings if f.status == "partially_compliant")),
        ("Non-Compliant", sum(1 for f in findings if f.status == "non_compliant")),
        ("Not Applicable", sum(1 for f in findings if f.status == "not_applicable")),
        ("Not Reviewed", sum(1 for f in findings if f.status == "not_reviewed")),
        ("", ""),
        ("── NIST SP 800-53A Formal Determinations ──", ""),
        ("Satisfied", satisfied),
        ("Other Than Satisfied", ots),
        ("Not Applicable", sum(1 for f in findings if f.status == "not_applicable")),
        ("", ""),
        ("── Assessment Quality Flags ──", ""),
        ("Synthesized Narratives", synthesized_count),
        ("AI Dissent Notes", challenged_count),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        cell = ws.cell(row=i, column=1, value=label)
        cell.font = Font(bold=("──" in str(label)))
        ws.cell(row=i, column=2, value=str(value) if value != "" else "")


def _write_control_matrix(ws, findings: list[ControlFinding]):
    headers = [
        "Control ID", "Family", "Title",
        "NIST Determination",        # Obs 8: formal 800-53A language for ATO package
        "Working Status",            # internal label kept for analyst reference
        "Implementation Statement",
        "Gaps", "Remediation Plan",
        "Reviewer Status", "Confidence",
        "Flags",                     # Obs 8: synthesized / challenge note indicators
    ]
    col_widths = [12, 8, 40, 22, 20, 60, 40, 50, 18, 12, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    _header_style(ws, 1, headers)

    for row_idx, f in enumerate(findings, start=2):
        nist_label = NIST_DETERMINATION_LABELS.get(f.status, "Not Reviewed")
        flags = []
        if getattr(f, "synthesized_narrative", False):
            flags.append("Synthesized Narrative")
        if getattr(f, "llm_challenge_note", None):
            flags.append("AI Dissent")

        values = [
            f.control_id,
            f.control_family,
            f.control_title,
            nist_label,
            STATUS_LABELS.get(f.status, f.status),
            f.implementation_statement or "",
            "\n".join(f.gaps) if f.gaps else "",
            f.remediation_plan or "",
            f.reviewer_status or "",
            f"{f.confidence_score:.2f}" if f.confidence_score is not None else "",
            " | ".join(flags) if flags else "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 4:  # NIST Determination column
                color = NIST_DETERMINATION_COLORS.get(nist_label, "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
            elif col_idx == 5:  # Working Status column (softer color)
                color = STATUS_COLORS.get(f.status, "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
            elif col_idx == 11 and flags:  # Flags column — amber highlight
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
        ws.row_dimensions[row_idx].height = 60

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_family_sheet(ws, family: str, findings: list[ControlFinding]):
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 18

    _header_style(ws, 1, [
        "Control ID", "Title", "NIST Determination",
        "Implementation Statement", "Remediation Plan", "Flags",
    ])
    for i, f in enumerate(findings, start=2):
        nist_label = NIST_DETERMINATION_LABELS.get(f.status, "Not Reviewed")
        flags = []
        if getattr(f, "synthesized_narrative", False):
            flags.append("Synthesized")
        if getattr(f, "llm_challenge_note", None):
            flags.append("Dissent")
        vals = [
            f.control_id,
            f.control_title,
            nist_label,
            f.implementation_statement or "",
            f.remediation_plan or "",
            " | ".join(flags) if flags else "",
        ]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if j == 3:
                color = NIST_DETERMINATION_COLORS.get(nist_label, "FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
            elif j == 6 and flags:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")
        ws.row_dimensions[i].height = 60
    ws.freeze_panes = "A2"

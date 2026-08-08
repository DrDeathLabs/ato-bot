"""Excel (.xlsx) parser using openpyxl."""
from __future__ import annotations

from pathlib import Path

from app.services.parsers.base import ParsedBlock, ParsedCell, ParsedDocument, ParsedPage


def parse_xlsx(file_path: str) -> ParsedDocument:
    try:
        import openpyxl
    except ImportError:
        return ParsedDocument(
            filename=Path(file_path).name,
            file_type="xlsx",
            error="openpyxl not installed",
        )

    doc = ParsedDocument(
        filename=Path(file_path).name,
        file_type="xlsx",
        parser_name="openpyxl",
        parser_version="sheet-table-v1",
    )
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        pages = []
        for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            page = ParsedPage(
                page_number=sheet_idx,
                content="",
                section_title=sheet_name,
                metadata={"sheet_name": sheet_name},
            )
            row_texts: list[str] = []
            header_cells: list[str | None] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                row_cells: list[ParsedCell] = []
                row_text_parts: list[str] = []
                non_empty = False
                for col_idx, cell in enumerate(row):
                    cell_text = "" if cell is None else str(cell).strip()
                    if row_idx == 0:
                        header = cell_text or None
                        header_cells.append(header)
                    else:
                        header = header_cells[col_idx] if col_idx < len(header_cells) else None
                    row_cells.append(
                        ParsedCell(
                            text=cell_text,
                            row_index=row_idx,
                            col_index=col_idx,
                            header=header,
                        )
                    )
                    if cell_text:
                        non_empty = True
                        row_text_parts.append(f"{header}: {cell_text}" if header and row_idx > 0 else cell_text)
                if not non_empty:
                    continue
                row_text = " | ".join(row_text_parts)
                page.blocks.append(
                    ParsedBlock(
                        block_id=f"{sheet_name}-row-{row_idx}",
                        block_type="table_row",
                        text=row_text,
                        heading_path=[sheet_name],
                        row_index=row_idx,
                        table_id=f"sheet-{sheet_idx}",
                        cells=row_cells,
                        metadata={"sheet_name": sheet_name},
                    )
                )
                row_texts.append(row_text)
            if row_texts:
                page.content = "\n".join(row_texts)
                pages.append(page)
        doc.pages = pages
        doc.metadata = {"sheet_count": len(pages)}
        doc.full_text = "\n\n".join(f"[Sheet: {p.section_title}]\n{p.content}" for p in pages)
        wb.close()
    except Exception as e:
        doc.error = str(e)
    return doc
